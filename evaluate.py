"""
evaluate.py — End-to-end evaluation: process a PDF with Claude, then
immediately compare the output against your hand-labeled CSVs.

This is the FIRST thing to run before using the tool on new PDFs.
At 24 pages per volume the entire run costs ~$0.50-1.00 and takes ~5 minutes.

Usage:
    # Full run against an existing PDF (recommended first step):
    python evaluate.py --pdf "Cartelera Teatral Madrileña 6 (1).pdf"

    # Quick smoke test — first 5 pages only:
    python evaluate.py --pdf "Cartelera Teatral Madrileña 6 (1).pdf" --pages 5

    # If you already extracted to Excel, skip re-extraction and just validate:
    python evaluate.py --skip-extraction --out evaluation_output.xlsx

API key is read from GEMINI_API_KEY env var, or pass --api-key.
"""

import argparse
import os
import sys
import logging
from pathlib import Path
from difflib import SequenceMatcher

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env")

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent

CSV_CONFIGS: dict[str, tuple[str, str]] = {
    "cruz": (
        str(SCRIPT_DIR / "Madrid Theater Database - Teatro de la Cruz and Teatro del Príncipe(Teatro de la Cruz) (3).csv"),
        "Teatro de la Cruz",
    ),
    "principe": (
        str(SCRIPT_DIR / "Madrid Theater Database - Teatro de la Cruz and Teatro del Príncipe(Teatro del Príncipe) (2).csv"),
        "Teatro del Príncipe",
    ),
    "incomplete": (
        str(SCRIPT_DIR / "Madrid Theater Database - Teatro de la Cruz and Teatro del Príncipe(Incomplete data (1750-1757)) (1).csv"),
        "Unknown",
    ),
}

TEXT_FIELDS   = ["Play 1 Title", "Author 1 Name", "Company Director"]
NUMERIC_FIELDS = ["Play 1 Box Office Receipts"]

# ── logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(SCRIPT_DIR / "extractor.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)

# ── Excel formatting (copied from main.py so evaluate.py is self-contained) ───

def apply_excel_formatting(xlsx_path: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2E4057")
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, min(ws.max_row + 1, 200))),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 42)
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)

# ── ground truth loader ───────────────────────────────────────────────────────

def load_ground_truth() -> pd.DataFrame:
    dfs = []
    for _key, (csv_path, theater_name) in CSV_CONFIGS.items():
        p = Path(csv_path)
        if not p.exists():
            logger.warning("CSV not found, skipping: %s", csv_path)
            continue
        df = pd.read_csv(p, encoding="latin-1")
        df = df.iloc[:, :16]          # drop trailing empty columns
        df.insert(0, "Theater", theater_name)
        dfs.append(df)
    if not dfs:
        raise FileNotFoundError("No ground truth CSVs found next to evaluate.py")
    return pd.concat(dfs, ignore_index=True)

# ── string similarity ─────────────────────────────────────────────────────────

def _sim(a: str, b: str) -> float:
    return SequenceMatcher(None, str(a).lower().strip(), str(b).lower().strip()).ratio()

# ── validation report ─────────────────────────────────────────────────────────

def validate_and_report(extracted_path: str) -> None:
    W = 64  # report width

    print("\n" + "═" * W)
    print("  EVALUATION REPORT")
    print("═" * W)

    extracted = pd.read_excel(extracted_path, dtype=str, keep_default_na=False)
    gt = load_ground_truth()

    print(f"\n  Extracted records : {len(extracted):,}")
    print(f"  Ground truth rows : {len(gt):,}")

    # ── 1. Theater distribution ───────────────────────────────────────────────
    print(f"\n{'─'*W}")
    print("  1. THEATER DISTRIBUTION (extracted)")
    print(f"{'─'*W}")
    if "Theater" in extracted.columns:
        for theater, cnt in extracted["Theater"].value_counts().items():
            pct = cnt / len(extracted) * 100
            bar = "█" * int(pct / 2.5)
            print(f"  {theater:<30}  {cnt:>5,}  {bar}")
    else:
        print("  (no Theater column in output)")

    # ── 2. Season/year coverage ───────────────────────────────────────────────
    print(f"\n{'─'*W}")
    print("  2. DATE RANGE COVERAGE")
    print(f"{'─'*W}")
    # Match on (Theater, Performance Date) so rows from Cruz and Príncipe
    # never get conflated — they routinely run different plays on the same day.
    def _norm_theater(s: pd.Series) -> pd.Series:
        return (s.fillna("").astype(str).str.strip().str.lower()
                 .str.replace("í", "i", regex=False))

    ext_keys_df = extracted[["Theater", "Performance Date"]].dropna()
    gt_keys_df  = gt[["Theater", "Performance Date"]].dropna()
    ext_keys_df = ext_keys_df.assign(
        _k_theater=_norm_theater(ext_keys_df["Theater"]),
        _k_date=ext_keys_df["Performance Date"].astype(str).str.strip(),
    )
    gt_keys_df = gt_keys_df.assign(
        _k_theater=_norm_theater(gt_keys_df["Theater"]),
        _k_date=gt_keys_df["Performance Date"].astype(str).str.strip(),
    )
    ext_dates = ext_keys_df["_k_date"]
    gt_dates  = gt_keys_df["_k_date"]

    # Try to extract years for a quick range summary
    import re
    ext_years = sorted(set(
        int(m.group()) for d in ext_dates if (m := re.search(r'\b(17|18)\d\d\b', d))
    ))
    gt_years = sorted(set(
        int(m.group()) for d in gt_dates if (m := re.search(r'\b(17|18)\d\d\b', d))
    ))

    if ext_years:
        print(f"  Extracted years   : {ext_years[0]} – {ext_years[-1]}")
    else:
        print("  Extracted years   : (could not parse)")
    if gt_years:
        print(f"  Ground truth years: {gt_years[0]} – {gt_years[-1]}")

    # Overlap — keyed on (theater, date) tuples
    ext_key_set = set(zip(ext_keys_df["_k_theater"], ext_keys_df["_k_date"]))
    gt_key_set  = set(zip(gt_keys_df["_k_theater"],  gt_keys_df["_k_date"]))
    matched_keys = ext_key_set & gt_key_set
    only_gt_keys = gt_key_set - ext_key_set
    only_ext_keys = ext_key_set - gt_key_set

    recall    = len(matched_keys) / len(gt_key_set)  * 100 if gt_key_set  else 0
    precision = len(matched_keys) / len(ext_key_set) * 100 if ext_key_set else 0

    print(f"\n  Matching (theater,date) : {len(matched_keys):,}  "
          f"(recall {recall:.1f}%  |  precision {precision:.1f}%)")
    print(f"  Only in GT              : {len(only_gt_keys):,}")
    print(f"  Only in extracted       : {len(only_ext_keys):,}")

    # Keep the simple flat-date sets around only for the "zero overlap" warning.
    matched = matched_keys

    if not matched:
        print(
            "\n  ⚠  ZERO DATE OVERLAP.\n"
            "  The PDF likely covers a different time period than the existing CSVs.\n"
            "  This means there's no way to auto-validate accuracy — manual spot-check\n"
            "  is required. Open evaluation_output.xlsx and compare a sample by eye."
        )
        _print_sample(extracted)
        _print_recommendations(has_overlap=False, text_scores={})
        return

    # ── 3. Field accuracy on matched dates ────────────────────────────────────
    print(f"\n{'─'*W}")
    print(f"  3. FIELD-LEVEL ACCURACY  ({len(matched):,} matching dates)")
    print(f"{'─'*W}")
    print(f"  {'Field':<32}  {'Similarity':>10}  {'Pairs':>6}  Visual")

    # Build (theater, date)-keyed lookups
    gt_ann = gt.assign(_k=list(zip(_norm_theater(gt["Theater"]),
                                    gt["Performance Date"].astype(str).str.strip())))
    ext_ann = extracted.assign(_k=list(zip(_norm_theater(extracted["Theater"]),
                                            extracted["Performance Date"].astype(str).str.strip())))
    gt_idx  = gt_ann[gt_ann["_k"].isin(matched_keys)].drop_duplicates("_k").set_index("_k")
    ext_idx = ext_ann[ext_ann["_k"].isin(matched_keys)].drop_duplicates("_k").set_index("_k")

    text_scores: dict[str, float] = {}

    for field in TEXT_FIELDS:
        if field not in gt.columns or field not in extracted.columns:
            print(f"  {field:<32}  {'(column missing)':>10}")
            continue

        scores: list[float] = []
        for date in matched:
            try:
                gv  = str(gt_idx.at[date,  field]) if date in gt_idx.index  else ""
                ev  = str(ext_idx.at[date, field]) if date in ext_idx.index else ""
                if gv.strip() and ev.strip() and gv != "nan" and ev != "nan":
                    scores.append(_sim(gv, ev))
            except Exception:
                continue

        if scores:
            avg = sum(scores) / len(scores)
            text_scores[field] = avg
            pct = int(avg * 20)
            bar = "█" * pct + "░" * (20 - pct)
            grade = "✓ GOOD" if avg >= 0.85 else ("△ OK" if avg >= 0.65 else "✗ POOR")
            print(f"  {field:<32}  {avg:>8.3f}    {len(scores):>5,}  {bar}  {grade}")
        else:
            print(f"  {field:<32}  {'(no pairs)':>10}")

    # ── 4. Numeric accuracy ───────────────────────────────────────────────────
    for field in NUMERIC_FIELDS:
        if field not in gt.columns or field not in extracted.columns:
            continue
        exact = total = 0
        for date in matched:
            try:
                gv = str(gt_idx.at[date,  field]).strip() if date in gt_idx.index  else ""
                ev = str(ext_idx.at[date, field]).strip() if date in ext_idx.index else ""
                if gv and ev and gv != "nan" and ev != "nan":
                    total += 1
                    if gv == ev:
                        exact += 1
            except Exception:
                continue
        if total:
            print(f"\n  {field:<32}  exact match: {exact}/{total} = {exact/total*100:.1f}%")

    # ── 5. Error examples ─────────────────────────────────────────────────────
    worst_field = min(text_scores, key=text_scores.get) if text_scores else None
    if worst_field and text_scores.get(worst_field, 1) < 0.85:
        print(f"\n{'─'*W}")
        print(f"  4. WORST-FIELD EXAMPLES  ({worst_field})")
        print(f"{'─'*W}")
        examples_shown = 0
        for key in sorted(matched):
            if examples_shown >= 5:
                break
            try:
                gv = str(gt_idx.at[key,  worst_field]) if key in gt_idx.index  else ""
                ev = str(ext_idx.at[key, worst_field]) if key in ext_idx.index else ""
                if gv.strip() and ev.strip() and gv != "nan" and ev != "nan":
                    sc = _sim(gv, ev)
                    if sc < 0.80:
                        theater_nm, date_str = key
                        print(f"  {theater_nm} · {date_str}")
                        print(f"    GT : {gv[:70]}")
                        print(f"    EXT: {ev[:70]}")
                        print(f"    sim: {sc:.3f}")
                        examples_shown += 1
            except Exception:
                continue

    _print_sample(extracted)
    _print_recommendations(has_overlap=True, text_scores=text_scores)


def _print_sample(df: pd.DataFrame, n: int = 4) -> None:
    cols = [c for c in ["Theater", "Performance Date", "Play 1 Title",
                         "Author 1 Name", "Company Director"] if c in df.columns]
    print(f"\n{'─'*64}")
    print(f"  5. SAMPLE EXTRACTED RECORDS (first {n})")
    print(f"{'─'*64}")
    print(df[cols].head(n).to_string(index=False))


def _print_recommendations(has_overlap: bool, text_scores: dict) -> None:
    print(f"\n{'─'*64}")
    print("  NEXT STEPS")
    print(f"{'─'*64}")
    if not has_overlap:
        print("  • Open evaluation_output.xlsx and spot-check 20 rows against the PDF.")
        print("  • If dates / titles look correct, the tool is working — dates just")
        print("    fall outside the range of the existing hand-labeled CSVs.")
        print("  • If dates / titles look wrong, re-examine the system prompt in")
        print("    extractor.py — the page layout understanding may need refinement.")
    else:
        worst = [f for f, s in text_scores.items() if s < 0.85]
        if not worst:
            print("  ✓ All fields scoring ≥ 0.85 — accuracy looks good.")
            print("  • Ready to run on new PDFs via: python main.py")
        else:
            print(f"  ✗ Low-scoring fields: {', '.join(worst)}")
            print("  • Review the error examples above — are they OCR issues,")
            print("    abbreviations, or structural misreads?")
            print("  • Tune SYSTEM_PROMPT in extractor.py, then re-run this script.")
            print("  • Do NOT move to new PDFs until these fields score ≥ 0.85.")
    print(f"\n  Full log: extractor.log\n{'═'*64}\n")


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract a Cartelera Teatral PDF and evaluate against CSVs."
    )
    parser.add_argument(
        "--pdf",
        default=None,
        help='Path to PDF, e.g. "Cartelera Teatral Madrileña 6 (1).pdf"',
    )
    parser.add_argument(
        "--pages",
        type=int,
        default=None,
        help="Process only the first N pages (omit for all pages)",
    )
    parser.add_argument(
        "--api-key",
        default=os.environ.get("OPENAI_API_KEY", ""),
        help="OpenAI API key (defaults to OPENAI_API_KEY env var)",
    )
    parser.add_argument(
        "--initial-theater", default="",
        help='Starting theater if the PDF opens mid-section, e.g. "Teatro del Príncipe"',
    )
    parser.add_argument(
        "--initial-director", default="",
        help='Starting company director if known, e.g. "María Hidalgo"',
    )
    parser.add_argument(
        "--initial-season", default="",
        help='Starting season year if known, e.g. "1757-1758"',
    )
    parser.add_argument(
        "--out",
        default=str(SCRIPT_DIR / "evaluation_output.xlsx"),
        help="Path to write the extracted Excel (default: evaluation_output.xlsx)",
    )
    parser.add_argument(
        "--skip-extraction",
        action="store_true",
        help="Skip extraction and just run validation on an existing --out file",
    )
    args = parser.parse_args()

    # ── Extraction phase ──────────────────────────────────────────────────────
    if not args.skip_extraction:
        if not args.pdf:
            parser.error("--pdf is required unless --skip-extraction is set")
        if not args.api_key:
            parser.error(
                "An OpenAI API key is required.\n"
                "Set OPENAI_API_KEY or pass --api-key ..."
            )

        # Import here so the script still works for --skip-extraction with no deps
        from extractor import load_few_shot_examples, process_pdf

        pdf_path = str(SCRIPT_DIR / args.pdf) if not Path(args.pdf).is_absolute() else args.pdf
        if not Path(pdf_path).exists():
            sys.exit(f"PDF not found: {pdf_path}")

        page_msg = f"first {args.pages}" if args.pages else "all"
        print(f"\nExtracting {page_msg} page(s) from: {Path(pdf_path).name}")
        print(f"Output will be saved to: {args.out}\n")

        few_shot = load_few_shot_examples(CSV_CONFIGS)
        logger.info("Loaded %d few-shot examples", len(few_shot))

        def _progress(current: int, total: int, msg: str = "") -> None:
            bar_len = 30
            filled = int(bar_len * current / max(total, 1))
            bar = "█" * filled + "░" * (bar_len - filled)
            print(f"\r  [{bar}] {current}/{total}  {msg:<40}", end="", flush=True)

        df = process_pdf(
            pdf_path=pdf_path,
            api_key=args.api_key,
            few_shot_examples=few_shot,
            progress_callback=_progress,
            checkpoint_path=args.out.replace(".xlsx", ".checkpoint.json"),
            pages_limit=args.pages,
            initial_theater=args.initial_theater,
            initial_director=args.initial_director,
            initial_season=args.initial_season,
        )
        print()  # newline after progress bar

        logger.info("Writing %d records to %s", len(df), args.out)
        df.to_excel(args.out, index=False, engine="openpyxl")
        apply_excel_formatting(args.out)
        print(f"\n  Extracted {len(df):,} records → {args.out}")

    # ── Validation phase ──────────────────────────────────────────────────────
    if not Path(args.out).exists():
        sys.exit(f"Output file not found: {args.out}  (run without --skip-extraction first)")

    validate_and_report(args.out)


if __name__ == "__main__":
    main()
