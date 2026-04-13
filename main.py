"""
main.py — GUI application for Cartelera Teatral PDF → Excel extraction.

Run with:  python main.py
"""

import logging
import os
import threading
from pathlib import Path

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env")

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extractor import load_few_shot_examples, process_pdf

# ── Logging setup ─────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
LOG_PATH = SCRIPT_DIR / "extractor.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

# ── Ground truth CSV configs ──────────────────────────────────────────────────
# These files live next to main.py and are used for few-shot prompting.

CSV_CONFIGS: dict[str, tuple[str, str]] = {
    "cruz": (
        str(SCRIPT_DIR / "Madrid Theater Database - Teatro de la Cruz and Teatro del Príncipe(Teatro de la Cruz) (3).csv"),
        "Teatro de la Cruz",
    ),
    "principe": (
        str(SCRIPT_DIR / "Madrid Theater Database - Teatro de la Cruz and Teatro del Príncipe(Teatro del Príncipe) (2).csv"),
        "Teatro del Príncipe",
    ),
    "incomplete": (
        str(SCRIPT_DIR / "Madrid Theater Database - Teatro de la Cruz and Teatro del Príncipe(Incomplete data (1750-1757)) (1).csv"),
        "Unknown",
    ),
}

# ── Excel formatting ──────────────────────────────────────────────────────────

def apply_excel_formatting(xlsx_path: str) -> None:
    """Apply header style, auto-column widths, and freeze the top row."""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2E4057")
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, header_cell in enumerate(ws[1], start=1):
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = center

        # Measure max content width in this column
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, min(ws.max_row + 1, 200))),  # sample first 200 rows
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 42)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    wb.save(xlsx_path)
    logger.info("Excel formatting applied: %s", xlsx_path)


# ── GUI ───────────────────────────────────────────────────────────────────────

class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Cartelera Teatral  —  PDF → Excel Extractor")
        self.root.geometry("660x500")
        self.root.resizable(False, False)

        self.pdf_path: str | None = None
        self.out_path: str | None = None

        self._build_ui()

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        H_PAD = {"padx": 24, "pady": 6}

        # Title bar
        title_frame = tk.Frame(self.root, bg="#2E4057", height=64)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame,
            text="Cartelera Teatral Madrileña",
            font=("Helvetica", 17, "bold"),
            bg="#2E4057", fg="white",
        ).pack(side="left", padx=24, pady=14)
        tk.Label(
            title_frame,
            text="PDF → Excel  |  Powered by Claude Vision",
            font=("Helvetica", 10),
            bg="#2E4057", fg="#AAC4E0",
        ).pack(side="left", pady=18)

        # ── API Key ──────────────────────────────────────────────────────────
        key_frame = tk.LabelFrame(self.root, text="  Gemini API Key  ", padx=10, pady=8)
        key_frame.pack(fill="x", **H_PAD)

        key_row = tk.Frame(key_frame)
        key_row.pack(fill="x")

        self.api_key_var = tk.StringVar(value=os.environ.get("GEMINI_API_KEY", ""))
        self._api_entry = tk.Entry(
            key_row, textvariable=self.api_key_var, show="•",
            font=("Courier", 11), relief="flat",
            bg="#F5F5F5", insertbackground="#333",
        )
        self._api_entry.pack(side="left", fill="x", expand=True, ipady=4)

        self._show_key = False
        self._eye_btn = tk.Button(
            key_row, text="👁", font=("Helvetica", 11),
            relief="flat", bg="#F5F5F5", cursor="hand2",
            command=self._toggle_key_visibility,
        )
        self._eye_btn.pack(side="right", padx=(4, 0))

        if self.api_key_var.get():
            tk.Label(key_frame, text="✓  Pre-filled from GEMINI_API_KEY environment variable",
                     fg="#388E3C", font=("Helvetica", 9)).pack(anchor="w", pady=(2, 0))

        # ── PDF input ────────────────────────────────────────────────────────
        pdf_frame = tk.LabelFrame(self.root, text="  Input PDF  ", padx=10, pady=8)
        pdf_frame.pack(fill="x", **H_PAD)

        pdf_row = tk.Frame(pdf_frame)
        pdf_row.pack(fill="x")

        self._pdf_label = tk.Label(
            pdf_row, text="No file selected",
            fg="#999999", font=("Helvetica", 10), anchor="w",
        )
        self._pdf_label.pack(side="left", fill="x", expand=True)
        tk.Button(
            pdf_row, text="Browse…", width=10, cursor="hand2",
            command=self._select_pdf,
        ).pack(side="right")

        # ── Output ───────────────────────────────────────────────────────────
        out_frame = tk.LabelFrame(self.root, text="  Output Excel File  ", padx=10, pady=8)
        out_frame.pack(fill="x", **H_PAD)

        out_row = tk.Frame(out_frame)
        out_row.pack(fill="x")

        self._out_label = tk.Label(
            out_row, text="Auto-set when PDF is selected",
            fg="#999999", font=("Helvetica", 10), anchor="w",
        )
        self._out_label.pack(side="left", fill="x", expand=True)
        tk.Button(
            out_row, text="Browse…", width=10, cursor="hand2",
            command=self._select_output,
        ).pack(side="right")

        # ── Progress ─────────────────────────────────────────────────────────
        prog_frame = tk.Frame(self.root)
        prog_frame.pack(fill="x", padx=24, pady=(12, 0))

        self._status_var = tk.StringVar(value="Ready — select a PDF to begin.")
        tk.Label(
            prog_frame, textvariable=self._status_var,
            font=("Helvetica", 10), fg="#444444", anchor="w",
        ).pack(fill="x")

        self._progress_var = tk.DoubleVar()
        self._progress_bar = ttk.Progressbar(
            prog_frame, variable=self._progress_var, maximum=100, length=612,
        )
        self._progress_bar.pack(fill="x", pady=(4, 0))

        # ── Extract button ────────────────────────────────────────────────────
        self._extract_btn = tk.Button(
            self.root,
            text="⚡  Extract to Excel",
            command=self._run_extraction,
            bg="#2E7D32", fg="white",
            font=("Helvetica", 13, "bold"),
            relief="flat", padx=28, pady=10,
            cursor="hand2", activebackground="#1B5E20", activeforeground="white",
        )
        self._extract_btn.pack(pady=22)

        tk.Label(
            self.root,
            text="Logs are written to  extractor.log  in the same folder.",
            font=("Helvetica", 9), fg="#AAAAAA",
        ).pack()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _toggle_key_visibility(self) -> None:
        self._show_key = not self._show_key
        self._api_entry.config(show="" if self._show_key else "•")

    def _select_pdf(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Cartelera Teatral PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if not path:
            return
        self.pdf_path = path
        self._pdf_label.config(text=Path(path).name, fg="#111111")

        # Auto-suggest output next to PDF
        default_out = str(Path(path).with_suffix(".xlsx"))
        self.out_path = default_out
        self._out_label.config(text=Path(default_out).name, fg="#111111")
        self._status_var.set(f"Selected: {Path(path).name}")

    def _select_output(self) -> None:
        initial = Path(self.out_path).name if self.out_path else "output.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save Excel Output As",
            initialfile=initial,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.out_path = path
            self._out_label.config(text=Path(path).name, fg="#111111")

    # ── Extraction workflow ───────────────────────────────────────────────────

    def _run_extraction(self) -> None:
        if not self.pdf_path:
            messagebox.showerror("Missing Input", "Please select a PDF file first.")
            return
        if not self.out_path:
            messagebox.showerror("Missing Output", "Please choose an output file path.")
            return

        api_key = self.api_key_var.get().strip()
        if not api_key:
            messagebox.showerror(
                "Missing API Key",
                "Please enter a valid Google Gemini API key "
                "(get one at https://aistudio.google.com/apikey).",
            )
            return

        self._extract_btn.config(state="disabled")
        self._progress_var.set(0)
        self._status_var.set("Loading few-shot examples from CSV files…")

        thread = threading.Thread(target=self._worker, args=(api_key,), daemon=True)
        thread.start()

    def _worker(self, api_key: str) -> None:
        checkpoint = str(Path(self.out_path).with_suffix(".checkpoint.json"))
        try:
            few_shot = load_few_shot_examples(CSV_CONFIGS)
            logger.info("Loaded %d few-shot examples", len(few_shot))

            df = process_pdf(
                pdf_path=self.pdf_path,
                api_key=api_key,
                few_shot_examples=few_shot,
                progress_callback=self._update_progress,
                checkpoint_path=checkpoint,
            )

            self._set_status("Writing Excel file…")
            df.to_excel(self.out_path, index=False, engine="openpyxl")

            self._set_status("Applying formatting…")
            apply_excel_formatting(self.out_path)

            # Delete checkpoint on clean finish
            cp = Path(checkpoint)
            if cp.exists():
                cp.unlink()

            self.root.after(0, self._on_success, len(df))

        except Exception as exc:
            logger.exception("Extraction failed")
            self.root.after(0, self._on_error, str(exc))

    def _update_progress(self, current: int, total: int, msg: str = "") -> None:
        pct = (current / max(total, 1)) * 100
        self.root.after(0, self._progress_var.set, pct)
        if msg:
            self.root.after(0, self._status_var.set, msg)

    def _set_status(self, msg: str) -> None:
        self.root.after(0, self._status_var.set, msg)

    def _on_success(self, record_count: int) -> None:
        self._extract_btn.config(state="normal")
        self._progress_var.set(100)
        self._status_var.set(f"Done — {record_count:,} records extracted successfully.")
        messagebox.showinfo(
            "Extraction Complete",
            f"Extracted {record_count:,} performance records.\n\n"
            f"Saved to:\n{self.out_path}",
        )

    def _on_error(self, error_msg: str) -> None:
        self._extract_btn.config(state="normal")
        self._status_var.set("Error — check extractor.log for details.")
        messagebox.showerror(
            "Extraction Failed",
            f"An error occurred:\n\n{error_msg}\n\n"
            "Check  extractor.log  for the full traceback.",
        )


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
